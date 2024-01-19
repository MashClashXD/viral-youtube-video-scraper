from flask import Flask, render_template, request, send_file
from googleapiclient.discovery import build
import requests
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)

def make_sheet(df, video_link, views, subscriber_count, title):
    new_row = {'Video Link': video_link, 'Subscribers': subscriber_count, 'Views': views, 'Title': title}
    df = df.append(new_row, ignore_index=True)
    df['Subscribers'] = df['Subscribers'].apply(lambda x: '{:,}'.format(x) if pd.notnull(x) and isinstance(x, (int, float)) else x)
    df['Views'] = df['Views'].apply(lambda x: '{:,}'.format(x) if pd.notnull(x) and isinstance(x, (int, float)) else x)
    df['Thumbnail'] = df['Video Link'].apply(lambda url: f'=IMAGE("https://img.youtube.com/vi/{url[url.find("v=") + 2:url.find("v=") + 13]}/maxresdefault.jpg", 1)')

    # Reorder columns with 'Thumbnail' as the second-to-last column
    df = df[['Video Link', 'Subscribers', 'Views', 'Thumbnail', 'Title']]

    return df

def get_subscriber_count(api_key, channel_id):
    url = f'https://www.googleapis.com/youtube/v3/channels?part=statistics&id={channel_id}&key={api_key}'
    response = requests.get(url)
    data = response.json()

    try:
        subscriber_count = data['items'][0]['statistics']['subscriberCount']
        return int(subscriber_count)
    except (KeyError, IndexError):
        return None

def get_days_from_time_filter(time_filter):
    if time_filter == '24 Hours':
        return 1
    elif time_filter == '1 Month':
        return 30
    elif time_filter == '6 Months':
        return 180
    elif time_filter == '1 Year':
        return 365
    else:
        return 0  # All Time

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/search', methods=['POST'])
def search_youtube():
    api_key = request.form.get('api_key')
    search_term = request.form.get('search_term')
    time_filter = request.form.get('time_filter')
    value_filter = request.form.get('value_filter')

    upload_date = (datetime.now() - timedelta(days=get_days_from_time_filter(time_filter))).isoformat('T') + 'Z'

    youtube = build('youtube', 'v3', developerKey=api_key)

    # Create an empty DataFrame to store the data
    df = pd.DataFrame(columns=['Video Link', 'Views', 'Subscribers'])

    search_response = youtube.search().list(
        q=search_term,
        type='video',
        part='id,snippet',
        order='viewCount',
        publishedAfter=upload_date,
        videoDuration='long',
        maxResults=10000
    ).execute()

    for search_result in search_response.get('items', []):
        video_id = search_result['id']['videoId']

        video_response = youtube.videos().list(
            id=video_id,
            part='snippet,statistics'
        ).execute()

        if 'snippet' in video_response['items'][0] and 'statistics' in video_response['items'][0]:
            snippet = video_response['items'][0]['snippet']
            statistics = video_response['items'][0]['statistics']

            channel_id = snippet['channelId']
            subscriber_count = get_subscriber_count(api_key, channel_id)
            title = snippet.get('title', '')

            views = int(statistics.get('viewCount', 0))
            video_link = f"https://www.youtube.com/watch?v={video_id}"

            if value_filter == 'Small (least restrictive)':
                if views > subscriber_count and views > 5000:
                    df = make_sheet(df, video_link, views, subscriber_count, title)
            elif value_filter == 'Medium (moderately restrictive)':
                if views > subscriber_count and subscriber_count < 1000000:
                    df = make_sheet(df, video_link, views, subscriber_count, title)
            elif value_filter == 'Large (most restrictive)':
                if views > subscriber_count and 1000 < subscriber_count < 100000 and views > 100000:
                    df = make_sheet(df, video_link, views, subscriber_count, title)

    # Create a new Excel Workbook
    workbook = Workbook()
    sheet = workbook.active

    # Define the header row
    header = ['Video Link', 'Subscribers', 'Views', 'Thumbnail', 'Title']
    for col_num, column_title in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write the data to the Excel file and adjust row heights
    for row_idx, row in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_num, value=cell_value)
            if isinstance(cell_value, str) and len(cell_value) > 50:
                sheet.row_dimensions[row_idx].height = 60  # Adjust row height for longer text

    # Apply styling to the 'Thumbnail' column (adjust column width)
    column_letter = get_column_letter(header.index('Thumbnail') + 1)
    column_dimension = sheet.column_dimensions[column_letter]
    column_dimension.width = 19.00

    # Change the path to save the Excel file to the /tmp directory
    excel_file_path = '/tmp/youtube_data.xlsx'

    # Save the Excel file
    workbook.save(excel_file_path)

    result = f"Data exported to 'youtube_data.xlsx'\n"
    file_generated = True
    return render_template('index.html', result=result, file_generated=file_generated)

@app.route('/download_excel')
def download_excel():
    try:
        # Specify the path to the generated Excel file in the /tmp directory
        excel_file_path = '/tmp/youtube_data.xlsx'

        return send_file(
            excel_file_path,
            as_attachment=True,
            download_name='youtube_data.xlsx'
        )
    except Exception as e:
        return str(e)

if __name__ == '__main__':
    app.run(debug=True)
